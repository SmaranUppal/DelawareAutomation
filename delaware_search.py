#!/usr/bin/env python3
"""
Delaware CourtConnect – Person/Case Search Automation
======================================================
Searches each name in NAMES, checks "Partial Last Name", sets a 60-day date
window, collects Case Number + Case Title + Filing Date from every result page,
deduplicates within the run and against any existing data in the Excel file,
then appends only new rows.

Requirements:
    pip install playwright openpyxl
    playwright install chromium
"""

import re
import time
from datetime import datetime, timedelta
from pathlib import Path

from playwright.sync_api import sync_playwright, Page
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CONFIG ────────────────────────────────────────────────────────────────────

NAMES = [
    # Business/entity names — full name goes in the Last Name / Company Name
    # field, First Name left blank.
    "J G W",
    "J.G. W",
    "J. G. W",
    "J.G.W",
    "JG W",
    "JGW",
    "J. G.W",
    "J GW",
    "Peachtree S",
    "Peach tree Settle",
    "DRB Cap",
    "Stone Street",
    "AA Ron I",
    "Abactor",
    "Abidole",
    "Adenna Med",
    "Adventura",
    "AGPI",
    "Aikman Structured Finance",
    "Annuity Transfers Ltd",
    "Apis Management",
    "Atlas Legal Funding III LP",
    "AXE Finance",
    "B.A.W.21",
    "B.R. Wright",
    "BHG Structured",
    "Bifco",
    "Blue Grape",
    "Catalina Structured Funding",
    "Concordis Group Limited",
    "Conrad Factoring",
    "Cornerstone Funding",
    "Fast Annuity S",
    "FL Assignments Corp",
    "G.D.T.R.F.B.",
    "G7 Crescenta",
    "Genex Capital Corp",
    "GJ 123",
    "Greenwood Funding",
    "Grier I",
    "Hakstol Group",
    "Hiddenview Ent, LLC",
    "JLC Capital Funding",
    "KN Direct Capital",
    "Lane Nimitz",
    "Lasko LLC",
    "Lasko, LLC",
    "Leaf 002 LLC",
    "Legere LLC",
    "Legere, LLC",
    "Lottery Funding",
    "M McDougall LLC",
    "M McDougall, LLC",
    "Majestic Funding",
    "Mic-Bry8",
    "Olive Branch Funding",
    "Palermo Group",
    "Palm Green Closing",
    "Palm Harbor",
    "Passira Mal",
    "Patriot Settlement",
    "QLS Funding",
    "Reliance Funding",
    "Rocorp Corporation",
    "RSL Funding",
    "Savannah Settlements",
    "Sempra Finance",
    "Seneca Originations",
    "SeneOne LLC",
    "Settlement Capital Corp",
    "Settlement Status",
    "Somerton LLC",
    "Somerton, LLC",
    "Stratcap Investments",
    "Stratton Asset",
    "Structured Asset",
    "TKD LLC",
    "TKD, LLC",
    "TRM V LLC",
    "TRM V, LLC",
    "Tybenz LLC",
    "Tybenz, LLC",
    "Uber Funding",
    "Vintage Equity Group",
    "Wepaymore Funding",
    "Zakho Way",
    "GREAT PLAINS MANAGEMENT CORPORATION",
    "RD FITZ LLC",
    "RD FITZ, LLC",
    "GA OFF LLC",
    "GA OFF, LLC",
    "Assured Management Corporation",
    "BENTZEN F",
]

OUTPUT_FILE = Path("delaware_cases.xlsx")
HEADLESS    = False   # set True to run without a visible browser window

SEARCH_URL  = (
    "https://courtconnect.courts.delaware.gov/cc/cconnect/"
    "ck_public_qry_cpty.cp_personcase_setup_idx"
)

MONTHS = ["JAN","FEB","MAR","APR","MAY","JUN",
          "JUL","AUG","SEP","OCT","NOV","DEC"]

def fmt_date(d: datetime) -> str:
    return f"{d.day:02d}-{MONTHS[d.month-1]}-{d.year}"

TODAY = datetime.today()
START = TODAY - timedelta(days=60)

# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────

SHEET      = "Cases"
HDR_FILL   = PatternFill("solid", start_color="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CELL_FONT  = Font(name="Arial", size=10)
EVEN_FILL  = PatternFill("solid", start_color="DCE6F1")
ODD_FILL   = PatternFill("solid", start_color="FFFFFF")
CTR        = Alignment(horizontal="center", vertical="center")
LFT        = Alignment(horizontal="left",   vertical="center")
_side      = Side(style="thin", color="BBBBBB")
BDR        = Border(left=_side, right=_side, top=_side, bottom=_side)
HEADERS    = ["Case Number", "Case Title", "Filing Date"]


def load_existing(path: Path) -> set[str]:
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path)
    if SHEET not in wb.sheetnames:
        return set()
    ws = wb[SHEET]
    return {
        str(row[0]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0]
    }


def init_workbook(path: Path):
    if path.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(1, col, h)
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CTR, BDR
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A2"
    wb.save(path)


def save_cases(path: Path, cases: list[dict]):
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET]
    start_row = ws.max_row + 1
    if start_row == 2 and ws.cell(2, 1).value is None:
        start_row = 2
    for i, case in enumerate(cases):
        row  = start_row + i
        fill = EVEN_FILL if row % 2 == 0 else ODD_FILL
        for col, val, align in [
            (1, case["case_number"], CTR),
            (2, case["case_title"],  LFT),
            (3, case["filing_date"], CTR),
        ]:
            c = ws.cell(row, col, val)
            c.font, c.alignment, c.border, c.fill = CELL_FONT, align, BDR, fill
    wb.save(path)


# ── SCRAPING ──────────────────────────────────────────────────────────────────

CASE_RE = re.compile(
    r"Case:\s+(\S+)\s+(.+)",   # group 1 = case number, group 2 = title
    re.IGNORECASE
)


def get_results_frame(page: Page):
    """Find the frame holding search results, robust to Playwright's
    frame-list ordering after a frameset reload. Prefers the frame named
    'Big'; falls back to scanning all frames for one containing case rows
    or the 'No records found' message.
    """
    # Preferred: frame explicitly named "Big" (matches the search form frame)
    for f in page.frames:
        if f.name == "Big":
            return f

    # Fallback: scan every frame for result markers
    for f in page.frames:
        try:
            txt = f.locator("body").inner_text(timeout=2000)
        except Exception:
            continue
        if "no records found" in txt.lower() or "Case:" in txt or "Name/Corporation" in txt:
            return f

    # Last resort: largest non-main frame, else main frame
    candidates = [f for f in page.frames if f != page.main_frame]
    if candidates:
        return max(candidates, key=lambda f: len(f.url or ""))
    return page.main_frame


def parse_rows(page: Page) -> list[dict]:
    """Extract all case rows from the current results frame."""
    cases = []
    frame = get_results_frame(page)

    rows = frame.query_selector_all("tr")
    for row in rows:
        cells = row.query_selector_all("td")
        if len(cells) < 6:
            continue

        addr_cell  = cells[2].inner_text()
        filing_raw = cells[5].inner_text().strip()

        # Case number is the hyperlink text inside the address cell
        link = cells[2].query_selector("a")
        if not link:
            continue
        case_num = link.inner_text().strip()

        # Title follows the case number on the "Case:" line
        match = CASE_RE.search(addr_cell)
        case_title = match.group(2).strip() if match else ""

        if case_num:
            cases.append({
                "case_number": case_num,
                "case_title":  case_title,
                "filing_date": filing_raw,
            })

    return cases


def search_name(page: Page, name: str) -> list[dict]:
    """Fill the search form, submit, and collect all paginated results.

    `name` is a business/entity name and goes entirely in the
    Last Name / Company Name field; First Name is left blank.
    """
    page.goto(SEARCH_URL, wait_until="networkidle", timeout=30_000)

    # The form is inside a frame called "Big"
    frame = next(f for f in page.frames if f.name == "Big")

    frame.fill('input[name="last_name"]',  name)
    frame.fill('input[name="first_name"]', "")

    # Ensure "Partial Last Name" is checked
    partial = frame.locator('input[name="partial_ind"]')
    if not partial.is_checked():
        partial.check()

    frame.fill('input[name="begin_date"]', fmt_date(START))
    frame.fill('input[name="end_date"]',   fmt_date(TODAY))

    frame.locator('input[type="submit"]').click()
    page.wait_for_load_state("networkidle", timeout=20_000)
    time.sleep(0.6)  # let the inner results frame finish rendering

    # Check for "no records"
    result_frame = get_results_frame(page)
    body_text = result_frame.locator("body").inner_text()
    if "no records found" in body_text.lower():
        print("    → No records found")
        return []

    all_cases: list[dict] = []
    pg = 1

    while True:
        rows = parse_rows(page)
        print(f"    Page {pg}: {len(rows)} rows")
        all_cases.extend(rows)

        # Look for Next link in result frame
        result_frame = get_results_frame(page)
        next_link = result_frame.locator("a", has_text=re.compile(r"next", re.IGNORECASE))
        if next_link.count() == 0:
            break

        next_link.first.click()
        page.wait_for_load_state("networkidle", timeout=20_000)
        pg += 1
        time.sleep(0.4)

    return all_cases


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Delaware CourtConnect Search")
    print(f"Date range : {fmt_date(START)} → {fmt_date(TODAY)}")
    print(f"Output file: {OUTPUT_FILE.resolve()}")
    print("=" * 60)

    init_workbook(OUTPUT_FILE)
    existing = load_existing(OUTPUT_FILE)
    print(f"Existing cases in file: {len(existing)}\n")

    new_cases:  list[dict] = []
    seen_run:   set[str]   = set()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        ctx     = browser.new_context()
        page    = ctx.new_page()

        for name in NAMES:
            print(f"→ Searching: {name}")
            results = search_name(page, name)

            for case in results:
                cn = case["case_number"]
                if cn in seen_run:
                    continue
                seen_run.add(cn)
                if cn in existing:
                    print(f"    SKIP (already saved): {cn}")
                    continue
                new_cases.append(case)

            time.sleep(1)

        browser.close()

    print(f"\n{'=' * 60}")
    print(f"New unique cases to add: {len(new_cases)}")

    if new_cases:
        save_cases(OUTPUT_FILE, new_cases)
        for c in new_cases:
            print(f"  [{c['case_number']}]  {c['case_title']}  |  {c['filing_date']}")
    else:
        print("Nothing new to add.")

    print(f"\nSaved → {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()